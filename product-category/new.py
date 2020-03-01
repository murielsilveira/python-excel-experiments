import os
from openpyxl import Workbook
current_path = os.path.dirname(os.path.realpath(__file__))


def fill_data(data, curr_sheet):
    for i, row in enumerate(data, start=1):
        for j, val in enumerate(row, start=1):
            curr_sheet.cell(i, j, value=val)


cat1_data = [
    ["Category", "Color", "Size"],
    ["CAT1", "Red", "120"],
    ["CAT1", "Blue", "30"],
    ["CAT1", "Green", "45"],
    ["CAT1", "Papaya Whip", "60"],
]

cat2_data = [
    ["Category", "Color", "Size"],
    ["CAT2", "Black", "10"],
    ["CAT2", "Grey", "70"],
    ["CAT2", "Pink", "30"],
]

product_data = [
    ["Product", "Category"],
    ["PROD1", "CAT1"],
    ["PROD2", "CAT2"],
    ["PROD3", "CAT1"],
    ["PROD4", "CAT2"],
    ["PROD5", "CAT1"],
    ["PROD6", "CAT2"],
    ["PROD7", "CAT2"],
    ["PROD8", "CAT1"],
    ["PROD9", "CAT2"],
    ["PROD10", "CAT1"],
]

workbook = Workbook()

cat1_sheet = workbook.active
cat1_sheet.title = "CAT1"
fill_data(cat1_data, cat1_sheet)

cat2_sheet = workbook.create_sheet("CAT2")
fill_data(cat2_data, cat2_sheet)

product_sheet = workbook.create_sheet("PRODUCTS")
fill_data(product_data, product_sheet)

workbook.save(filename=current_path + "/data/data.xlsx")
