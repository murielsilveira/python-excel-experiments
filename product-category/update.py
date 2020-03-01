import os
from openpyxl import load_workbook
current_path = os.path.dirname(os.path.realpath(__file__))

workbook = load_workbook(filename=current_path + "/data/data.xlsx")
product_sheet = workbook['PRODUCTS']
result_sheet = workbook.create_sheet("RESULT")

row_offset = 1
for product, category in product_sheet.iter_rows(min_row=2, values_only=True):
    cat_sheet = workbook[category]
    rows_added = 0

    if row_offset == 1:
        result_sheet.cell(1, 1, product_sheet["A1"].value)
        for col_index, (value,) in enumerate(cat_sheet.iter_cols(min_row=1, max_row=1, values_only=True), start=2):
            result_sheet.cell(1, col_index, value)

    for row_index, category_row in enumerate(cat_sheet.iter_rows(min_row=2, values_only=True), start=1):
        rows_added += 1
        result_sheet.cell(row_index + row_offset, 1, product)
        for col_index, value in enumerate(category_row, start=2):
            result_sheet.cell(row_index + row_offset, col_index, value)

    row_offset += rows_added

workbook.save(filename=current_path + "/data/data_result.xlsx")
